"""
Excel 导出模块 — 将提取结果汇总导出为 Excel 文件
"""

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def results_to_excel(
    json_path: str,
    output_path: str,
    template_fields: dict[str, list[str]] | None = None,
) -> str:
    """
    将 JSON 结果文件转换为 Excel

    Args:
        json_path: 输入的 JSON 文件路径
        output_path: 输出的 Excel 文件路径
        template_fields: 可选，字段顺序模板 {项目类型: [字段列表]}

    Returns:
        输出的 Excel 文件路径
    """
    with open(json_path, "r", encoding="utf-8") as f:
        results = json.load(f)

    if not results:
        raise ValueError("JSON 文件为空，没有可导出的数据")

    wb = Workbook()
    ws = wb.active
    ws.title = "项目汇总"

    # ---- 收集所有字段 ----
    all_fields = ["项目类型"]
    field_set = set(all_fields)
    for item in results:
        for k in item.keys():
            if k not in field_set:
                all_fields.append(k)
                field_set.add(k)

    # ---- 样式 ----
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ---- 写表头 ----
    for col_idx, field in enumerate(all_fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ---- 写数据 ----
    for row_idx, item in enumerate(results, 2):
        # 第一列：项目类型
        cell = ws.cell(row=row_idx, column=1, value=item.get("项目类型", ""))
        cell.alignment = cell_alignment
        cell.border = thin_border

        for col_idx, field in enumerate(all_fields[1:], 2):
            value = item.get(field, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False, indent=2)
            elif value is None:
                value = ""
            else:
                value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # ---- 列宽自适应（最大 50 字符） ----
    for col_idx, field in enumerate(all_fields, 1):
        max_len = len(field) * 2  # 中文字符占位更宽
        for row_idx in range(2, len(results) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                # 中文字符算 2 个宽度
                char_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
                max_len = max(max_len, min(char_len, 50))
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max_len + 2

    # ---- 冻结首行 ----
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def append_results_to_excel(
    input_excel: str,
    results: list[dict],
    output_path: str,
) -> str:
    """
    在原 Excel 后面追加提取结果列

    对于每个项目：
    - 匹配 original 的「完整路径」列（或「项目路径」列）
    - 在右侧新增列：提取字段 + 置信度 + 是否一致

    Args:
        input_excel: 原始 Excel 文件
        results: 提取结果列表
        output_path: 输出 Excel 文件路径

    Returns:
        输出路径
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # 构建结果索引 {完整路径: 结果}
    result_index: dict[str, dict] = {}
    for r in results:
        path_key = r.get("_path", "").replace("\\", "/").lower().strip()
        result_index[path_key] = r

    # 加载原始 Excel
    input_path = Path(input_excel)
    wb = load_workbook(input_path)
    ws = wb.active

    if ws is None:
        raise ValueError("Excel 中没有工作表")

    # 读表头，找到「完整路径」列
    headers = []
    path_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        val = str(cell.value).strip() if cell.value else ""
        headers.append(val)
        if val in ("完整路径", "项目路径", "文件夹路径"):
            path_col = col_idx

    if path_col is None:
        # 尝试找第一列包含"路径"的
        for i, h in enumerate(headers, 1):
            if "路径" in h:
                path_col = i
                break

    if path_col is None:
        raise ValueError(f"原始 Excel 中找不到路径列。表头: {headers}")

    # 找原始数据真正的最后一列（表头有值的最后一列）
    max_col = ws.max_column
    # 从最后一列往前扫，找表头有值的列
    while max_col > 0:
        header_val = ws.cell(row=1, column=max_col).value
        if header_val:
            break
        max_col -= 1

    # 如果 max_col 过大（原始表拖了 16384 列），回退到已知列之后
    orig_header_count = len([h for h in headers if h])
    if max_col > orig_header_count + 50:
        max_col = orig_header_count

    # 收集原始 Excel 中所有存在的类型（项目小类列）
    type_col = None
    for i, h in enumerate(headers, 1):
        if h in ("项目小类", "项目类型"):
            type_col = i
            break

    # 新增列的表头
    new_headers = [
        "企业名称_提取",
        "企业地址_提取",
        "年份_提取",
        "整治类型_提取",
        "排水许可证编号_提取",
        "置信度",
        "与人工分类是否一致",
        "需人工复核",
        "验证通过字段",
    ]

    # 样式
    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 浅黄
    err_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # 浅红

    # 写新表头
    for offset, h in enumerate(new_headers):
        col = max_col + 1 + offset
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 逐行匹配并写入
    matched = 0
    unmatched = 0
    for row_idx in range(2, ws.max_row + 1):
        path_cell = ws.cell(row=row_idx, column=path_col)
        if not path_cell.value:
            continue
        path_val = str(path_cell.value).replace("\\", "/").lower().strip()

        result = result_index.get(path_val)
        if result is None:
            unmatched += 1
            continue

        matched += 1
        extracted = result.get("提取结果", {})
        confidence = result.get("置信度评估", {}) or {}
        verify = result.get("验证", {}) or {}

        # 判断是否一致：提取的整治类型 vs Excel 中的项目小类
        extracted_type = extracted.get("整治类型", "")
        excel_type = ""
        if type_col:
            excel_type = str(ws.cell(row=row_idx, column=type_col).value or "").strip()

        match = _type_match(extracted_type, excel_type)
        need_review = confidence.get("需人工复核", False) or (not match)
        verified_count = confidence.get("已验证字段数", 0)
        total_verified = verified_count + confidence.get("未找到原文字段数", 0)

        values = [
            extracted.get("企业名称", ""),
            extracted.get("企业地址", ""),
            extracted.get("年份", ""),
            extracted_type,
            extracted.get("排水许可证编号", ""),
            confidence.get("等级", ""),
            "一致" if match else ("不一致" if extracted_type else "未识别"),
            "是" if need_review else "否",
            f"{verified_count}/{total_verified}" if total_verified else "-",
        ]

        for offset, val in enumerate(values):
            col = max_col + 1 + offset
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            # 不一致的行标黄
            if not match and offset == 6:
                cell.fill = warn_fill
            # 需复核的行标浅红
            if need_review and offset == 7:
                cell.fill = err_fill if need_review else None

    # 调整列宽
    for offset in range(len(new_headers)):
        col = max_col + 1 + offset
        max_len = len(new_headers[offset]) * 2
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col).value
            if val:
                char_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
                max_len = max(max_len, min(char_len, 40))
        col_letter = _col_letter(col)
        ws.column_dimensions[col_letter].width = max_len + 2

    output_path = Path(output_path)
    wb.save(output_path)
    print(f"   匹配 {matched} 行，未匹配 {unmatched} 行")
    return str(output_path)


def _type_match(extracted: str, excel: str) -> bool:
    """判断提取的整治类型与 Excel 项目小类是否一致"""
    if not extracted or not excel:
        return False
    e = extracted.replace("整治", "").replace("工业", "").replace("企业", "").replace(" ", "")
    x = excel.replace("整治", "").replace("工业", "").replace("企业", "").replace(" ", "")
    # 核心词匹配
    if "雨污分流" in e and "雨污分流" in x:
        return True
    if "废水接纳" in e and "废水接纳" in x:
        return True
    return False


def _col_has_data(ws, col: int, rows_to_check: int = 10) -> bool:
    """检查某列前N行是否有数据"""
    for row in range(1, min(rows_to_check + 1, ws.max_row + 1)):
        if ws.cell(row=row, column=col).value is not None:
            return True
    return False


def _col_letter(col: int) -> str:
    """列号转字母（1→A, 27→AA）"""
    result = ""
    while col > 0:
        col -= 1
        result = chr(65 + col % 26) + result
        col //= 26
    return result


def merge_json_results(results: list[dict], output_path: str) -> str:
    """
    将多个项目提取结果合并为一个 JSON 文件
    
    Args:
        results: 项目结果列表
        output_path: 输出的 JSON 文件路径
        
    Returns:
        输出的 JSON 文件路径
    """
    output_path = Path(output_path).resolve()
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    return str(output_path)
