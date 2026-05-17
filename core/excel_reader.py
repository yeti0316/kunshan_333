"""
Excel 读取模块 — 读取项目清单 Excel，返回待处理的项目列表
"""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_project_list(excel_path: str) -> list[dict[str, Any]]:
    """
    读取项目清单 Excel，返回项目列表
    
    Excel 格式要求：
    - 第一行为表头
    - 必须包含列：项目路径（或文件夹路径）、项目类型
    - 其他列可选，会原样保留
    
    Args:
        excel_path: Excel 文件路径
        
    Returns:
        项目列表，每项包含 path, type 及其他列信息
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError("Excel 文件中没有工作表")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件为空")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    # 找到必要的列索引
    path_col = _find_col_index(headers, [
        "完整路径", "项目路径", "文件夹路径", "路径", "project_path"
    ])
    type_col = _find_col_index(headers, [
        "项目小类", "项目类型", "类型", "project_type"
    ])

    if path_col is None:
        raise ValueError(
            f"Excel 中找不到「项目路径」列。当前表头: {headers}\n"
            f"请确保有一列名为「项目路径」或「文件夹路径」"
        )
    if type_col is None:
        raise ValueError(
            f"Excel 中找不到「项目类型」列。当前表头: {headers}\n"
            f"请确保有一列名为「项目类型」"
        )

    projects = []
    for row in data_rows:
        row_path = str(row[path_col]).strip() if row[path_col] else ""
        row_type = str(row[type_col]).strip() if row[type_col] else ""

        # 跳过空行
        if not row_path or not row_type:
            continue

        project = {
            "path": row_path,
            "type": row_type,
        }

        # 保留所有原始列
        for i, header in enumerate(headers):
            if i < len(row) and i not in (path_col, type_col):
                val = row[i]
                if val is not None:
                    project[header] = str(val).strip()

        projects.append(project)

    wb.close()

    if not projects:
        raise ValueError("Excel 中没有有效的项目数据")

    return projects


def _find_col_index(headers: list[str], candidates: list[str]) -> int | None:
    """在表头中查找第一个匹配的列索引"""
    for i, h in enumerate(headers):
        for c in candidates:
            if c == h or c in h:
                return i
    return None


def read_directory_as_project_list(root_dir: str) -> list[dict[str, Any]]:
    """
    当没有 Excel 清单时，直接把一个目录下的所有子文件夹作为项目列表
    项目类型需要从文件夹名判断，或统一为"未分类"
    
    Args:
        root_dir: 根目录
        
    Returns:
        项目列表
    """
    root = Path(root_dir)
    if not root.exists():
        raise FileNotFoundError(f"目录不存在: {root_dir}")

    projects = []
    for p in sorted(root.iterdir()):
        if p.is_dir() and not p.name.startswith("."):
            projects.append({
                "path": str(p),
                "type": "未分类",  # 需要后续匹配
                "folder_name": p.name,
            })

    if not projects:
        raise ValueError(f"目录 {root_dir} 下没有子文件夹")

    return projects
