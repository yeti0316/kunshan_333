from openpyxl import load_workbook

wb = load_workbook('多源项目信息提取器/output/工业企业整治结果_fix2.xlsx')
ws = wb.active

# 显示全部表头
print("全部表头:")
for c in range(1, 30):
    v = ws.cell(row=1, column=c).value
    if v:
        print(f"  列{c}: {v}")
print()

# 看第2行
print("第2行数据:")
for c in range(10, 25):
    h = ws.cell(row=1, column=c).value
    val = ws.cell(row=2, column=c).value
    print(f"  列{c} ({h}): {val}")
