"""检查 Excel 输出值"""
import json
from openpyxl import load_workbook

# 读 JSON 看第一条结果
with open('多源项目信息提取器/output/工业企业整治结果_v5.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

r = data[0]
print("JSON 中第一条:")
print(f"  _path: {r.get('_path')}")
print(f"  提取结果: {r.get('提取结果')}")
print()

# 读 Excel
wb = load_workbook('多源项目信息提取器/output/工业企业整治结果_v5.xlsx')
ws = wb.active

# 看表头最后几列
max_col = ws.max_column
print(f"Excel 共 {max_col} 列")
headers = []
for c in range(1, max_col + 1):
    headers.append(str(ws.cell(row=1, column=c).value or ''))
print(f"表头后10列: {headers[-10:]}")

# 看第2行（第一条数据）的后几列
print("\n第2行后10列:")
for c in range(max_col - 9, max_col + 1):
    val = ws.cell(row=2, column=c).value
    print(f"  列{c}({headers[c-1][:20]}): {val}")
